#Created by Arthur Jackson
#Rosetta Stone Requirement Matcher v0.3 Sept. 18, 2024

#Requirements from different documents may need to be matched in the case that each document has them written differently and out of order.
#This program will, for each requirement in column 1, parse through each requirement in column 2 to find its counterpart.
#This is done using Levenshtein's (word level) method, which counts the number of words you would need to change to match the phrases.


from openpyxl import load_workbook

# Function to calculate word-level Levenshtein distance between two strings
def word_level_edit_distance(str1, str2):
    words1 = str1.lower().split()
    words2 = str2.lower().split()
    
    # Get the length of both word lists
    len1 = len(words1)
    len2 = len(words2)
    
    # Create a matrix to store distances
    dp = [[0 for _ in range(len2 + 1)] for _ in range(len1 + 1)]

    # Initialize the matrix
    for i in range(len1 + 1):
        dp[i][0] = i
    for j in range(len2 + 1):
        dp[0][j] = j

    # Fill the matrix
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if words1[i - 1] == words2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                # Compute the minimum cost of deletion, insertion, or substitution
                dp[i][j] = min(dp[i - 1][j] + 1,  
                               dp[i][j - 1] + 1,  
                               dp[i - 1][j - 1] + 1)  

    
    return dp[len1][len2]

#wb_name = input("Enter name of xlsx file (inlude .xlsx)")
wb = load_workbook('updated_file_word_level2.xlsx')
ws = wb.active

# Iterate through column B (starting from row 1)
for row_b in range(1, ws.max_row + 1):
    value_b = ws[f'B{row_b}'].value  # Get the value from column B
    id_b = ws[f'A{row_b}'].value  # Get the ID from column A (left of column B)
    if value_b is None:  # Skip if the cell is empty
        continue

    best_match = None
    best_match_id = None
    lowest_edit_distance = float('inf')  # Set to a very large number initially

    # Iterate through column D to find the best match based on word-level edit distance
    for row_d in range(1, ws.max_row + 1):
        print("First Row: " + str(row_b) + " Second Row: " + str(row_d))
        value_d = ws[f'D{row_d}'].value  # Get the value from column D
        id_d = ws[f'C{row_d}'].value  # Get the ID from column C (left of column D)
        if value_d is None:  # Skip if the cell is empty
            continue
        
        # Calculate the word-level edit distance (case-insensitive comparison)
        edit_distance = word_level_edit_distance(str(value_b), str(value_d))
        
        # If the edit distance is lower than the current best match, update it
        if edit_distance < lowest_edit_distance:
            lowest_edit_distance = edit_distance
            best_match = value_d
            best_match_id = id_d  # Store the corresponding ID from column C

    # Place the value from B and its best match from D into columns G and H, and their IDs into F and I
    ws[f'F{row_b}'] = id_b  # Column F stores the ID from column A (left of B)
    ws[f'G{row_b}'] = value_b  # Column G stores the value from column B
    

    ws[f'H{row_b}'] = best_match if best_match else 'No match' 
    ws[f'I{row_b}'] = best_match_id if best_match_id else 'No ID' 


wb.save('WithIdTest.xlsx')

print("Process completed. The matches and IDs are stored in columns F, G, H, and I.")




#Did you know?:
#Rats cannot throw up